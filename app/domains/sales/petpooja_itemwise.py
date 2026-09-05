"""Reading a Petpooja "Item Wise: Sales Report" export.

D15 dismissed this report for bills — it is pre-aggregated and carries no
bill number — and that stands. What it DOES carry, and nothing else does, is
the menu's own taxonomy: every item printed under its category, with
Petpooja's item code beside it. That map is what turns the names-per-bill of
the Order Listing (D21) into per-bill attach rates, and it is why this
adapter exists.

**What this export actually looks like**, from AKIRA's own file:

    r0-r2   preamble — "Date:", "Name:", "Restaurant Name:"
    r3-r4   blank
    r5      header: Category, Item, Code, Sap Code, Qty., Total (₹)
    r6-r9   Total / Min. / Max. / Avg.
    r10+    items GROUPED by category: the category appears once, on the
            group's first row, and the rows below it carry a blank; each
            group ends with a "Sub Total" row

So the category is forward-filled, and "Sub Total" is a summary label like
the others. Item codes arrive as integers for most items and as a lowercase
name for a few ("cream cheese mushroo" — Petpooja truncated its own code),
so they are kept as text and never used as a key.

Pure. Bytes in, dataclasses out.
"""

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from app.domains.sales.petpooja import (
    SUMMARY_LABELS,
    UnreadableExport,
    parse_period,
    restaurant_in_preamble,
    to_paise,
)

ITEMWISE_ADAPTER_VERSION = "petpooja.itemwise.v1"

#: Identify the header row. "Category" + "Item" + "Qty." with no "Date" —
#: the Item Report: Day Wise has Item + Date + Qty. and no Category, and the
#: Category Wise report has Category but no Item.
ITEMWISE_HEADER_MARKERS = ("Category", "Item", "Qty.")


@dataclass(frozen=True)
class ParsedMenuItem:
    category: str
    #: Petpooja's printed name — the join key to sales_order_items.item_name
    #: and sales_item_days.item_name.
    item_name: str
    petpooja_code: str | None
    qty: float
    net_paise: int


@dataclass(frozen=True)
class ItemWiseResult:
    items: list[ParsedMenuItem]
    warnings: list[dict[str, Any]] = field(default_factory=list)
    period_start: date | None = None
    period_end: date | None = None
    restaurant: str | None = None
    adapter_version: str = ITEMWISE_ADAPTER_VERSION
    reported_qty: float | None = None

    @property
    def total_qty(self) -> float:
        return sum(i.qty for i in self.items)

    @property
    def categories(self) -> dict[str, list[str]]:
        out: dict[str, list[str]] = {}
        for i in self.items:
            out.setdefault(i.category, []).append(i.item_name)
        return out


def _qty(value: object) -> float:
    try:
        return float(Decimal(str(value).strip()))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"unreadable quantity: {value!r}") from exc


def parse_itemwise(data: bytes) -> ItemWiseResult:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise UnreadableExport(f"Not a readable .xlsx file: {exc}") from exc
    try:
        sheet = workbook[workbook.sheetnames[0]]
        raw = [tuple(r) for r in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    warnings: list[dict[str, Any]] = []
    period_start = period_end = None
    header: list[str] | None = None
    header_at = -1
    for index, row in enumerate(raw):
        label = str(row[0] or "").strip().rstrip(":").lower() if row else ""
        if label == "date" and len(row) > 1:
            period_start, period_end = parse_period(row[1])
        cells = {str(cell or "").strip() for cell in row}
        if all(m in cells for m in ITEMWISE_HEADER_MARKERS) and "Date" not in cells:
            header = [str(cell or "").strip() for cell in row]
            header_at = index
            break
    if header is None:
        raise UnreadableExport(
            "No header row containing 'Category', 'Item' and 'Qty.'. "
            "This does not look like a Petpooja Item Wise: Sales Report."
        )
    restaurant = restaurant_in_preamble(raw[:header_at])

    at = {name: header.index(name) for name in header if name}
    cat_col, item_col, qty_col = at["Category"], at["Item"], at["Qty."]
    code_col = at.get("Code")
    total_col = next((at[c] for c in at if c.startswith("Total")), None)

    def cell(row: tuple[Any, ...], pos: int | None) -> Any:
        return row[pos] if pos is not None and pos < len(row) else None

    items: list[ParsedMenuItem] = []
    seen: set[str] = set()
    current_category: str | None = None
    reported_qty: float | None = None
    for index, row in enumerate(raw[header_at + 1 :], start=header_at + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        first = str(cell(row, cat_col) or "").strip()
        if first.lower() in SUMMARY_LABELS:
            if first.lower() == "total" and reported_qty is None:
                try:
                    reported_qty = _qty(cell(row, qty_col))
                except ValueError:
                    warnings.append({"kind": "bad_total_row", "row": index})
            # "Sub Total" closes a group; the next data row names a new one.
            continue
        if first:
            current_category = first
        name = str(cell(row, item_col) or "").strip()
        if not name:
            continue
        if current_category is None:
            warnings.append({"kind": "item_before_any_category", "row": index, "item": name})
            continue
        if name in seen:
            warnings.append({"kind": "duplicate_item", "row": index, "item": name})
            continue
        seen.add(name)
        try:
            qty = _qty(cell(row, qty_col))
        except ValueError as exc:
            warnings.append({"kind": "bad_row", "row": index, "item": name, "detail": str(exc)})
            continue
        net = 0
        raw_total = cell(row, total_col)
        if raw_total is not None and str(raw_total).strip() != "":
            try:
                net = to_paise(raw_total)
            except (ValueError, ArithmeticError):
                warnings.append({"kind": "bad_amount", "row": index, "item": name})
        raw_code = cell(row, code_col)
        code = None
        if raw_code is not None and str(raw_code).strip() != "":
            code = str(int(raw_code)) if isinstance(raw_code, float) else str(raw_code).strip()
        items.append(
            ParsedMenuItem(
                category=current_category,
                item_name=name,
                petpooja_code=code,
                qty=qty,
                net_paise=net,
            )
        )

    if not items:
        raise UnreadableExport("The export contained no item rows.")
    return ItemWiseResult(
        items=items,
        warnings=warnings,
        period_start=period_start,
        period_end=period_end,
        restaurant=restaurant,
        reported_qty=reported_qty,
    )
