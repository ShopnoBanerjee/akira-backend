"""Reading a Petpooja Item Report: Day Wise export.

The one Petpooja export that carries TRUE UNITS: quantity and value per menu
item per day. The Order Listing (D21) has names without quantities; the Item
Wise report has quantities without days; this one has both, and it is what
makes theoretical consumption arithmetic instead of guesswork.

**What this export actually looks like**, established by reading a real file:

    r0-r2   preamble — "Date:", "Name:", "Restaurant Name:"
    r3-r4   blank
    r5      the header row: Item, Date, Qty., Total (₹)
    r6      a Total row — the same trap as the master's summary rows
    r7+     one row per (item, day), grouped by item

Two things worth stating plainly:

- **The date is Petpooja's own day grouping, taken verbatim.** No timestamps
  exist here to apply the 05:00 rule to. It lands in `report_date`, named so
  nobody mistakes it for a derived business_date.
- **Quantities can be fractional** (half portions exist) and arrive as
  strings ("5.0") or floats depending on the cell. Decimal handling matches
  the money rule: exact by construction.
"""

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from app.domains.sales.petpooja import (
    SUMMARY_LABELS,
    UnreadableExport,
    parse_period,
    to_paise,
)

ITEMDAYS_ADAPTER_VERSION = "petpooja.itemdays.v1"

#: Identify the header row. "Qty." is what separates this report from the
#: listing ("Order No." + "Items") and the master ("Invoice No.").
ITEMDAYS_HEADER_MARKERS = ("Item", "Date", "Qty.")


@dataclass(frozen=True)
class ParsedItemDay:
    item_name: str
    report_date: date
    qty: float
    net_paise: int


@dataclass(frozen=True)
class ItemDaysResult:
    rows: list[ParsedItemDay]
    warnings: list[dict[str, Any]] = field(default_factory=list)
    period_start: date | None = None
    period_end: date | None = None
    restaurant: str | None = None
    adapter_version: str = ITEMDAYS_ADAPTER_VERSION
    #: The export's own Total quantity, kept beside our sum so a
    #: disagreement is a visible pair, not a hand check.
    reported_qty: float | None = None

    @property
    def total_qty(self) -> float:
        return sum(r.qty for r in self.rows)

    @property
    def net_paise(self) -> int:
        return sum(r.net_paise for r in self.rows)


def _parse_report_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    found = re.match(r"(\d{4})-(\d{2})-(\d{2})", text)
    if not found:
        raise ValueError(f"unrecognised date: {value!r}")
    return date(int(found.group(1)), int(found.group(2)), int(found.group(3)))


def _parse_qty(value: object) -> float:
    try:
        return float(Decimal(str(value).strip()))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"unreadable quantity: {value!r}") from exc


def parse_itemdays(data: bytes) -> ItemDaysResult:
    """Read an Item Report: Day Wise export. Bytes in, (item, day, qty) out."""
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise UnreadableExport(f"Not a readable .xlsx file: {exc}") from exc

    try:
        sheet = workbook[workbook.sheetnames[0]]
        raw = [tuple(r) for r in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    warnings: list[dict[str, Any]] = []
    period_start = period_end = None
    restaurant = None
    header: list[str] | None = None
    header_at = -1

    for index, row in enumerate(raw):
        label = str(row[0] or "").strip().rstrip(":").lower() if row else ""
        if label == "date" and len(row) > 1:
            period_start, period_end = parse_period(row[1])
        elif label == "restaurant name" and len(row) > 1:
            restaurant = str(row[1] or "").strip() or None
        cells = {str(cell or "").strip() for cell in row}
        if all(marker in cells for marker in ITEMDAYS_HEADER_MARKERS):
            header = [str(cell or "").strip() for cell in row]
            header_at = index
            break

    if header is None:
        raise UnreadableExport(
            "No header row containing 'Item', 'Date' and 'Qty.'. "
            "This does not look like a Petpooja Item Report: Day Wise."
        )

    at = {name: header.index(name) for name in header if name}
    qty_col = at["Qty."]
    total_col = next((at[c] for c in at if c.startswith("Total")), None)

    rows: list[ParsedItemDay] = []
    seen: set[tuple[str, date]] = set()
    reported_qty: float | None = None

    for index, row in enumerate(raw[header_at + 1 :], start=header_at + 2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        first = str(row[0] or "").strip()
        if first.lower() in SUMMARY_LABELS:
            # The file's own Total row. Not data — but its quantity is the
            # export's own claim, kept for reconciliation.
            if first.lower() == "total" and reported_qty is None:
                try:
                    reported_qty = _parse_qty(row[qty_col])
                except (ValueError, IndexError):
                    warnings.append({"kind": "bad_total_row", "row": index})
            continue
        if not first:
            continue

        try:
            when = _parse_report_date(row[at["Date"]] if at["Date"] < len(row) else None)
            qty = _parse_qty(row[qty_col] if qty_col < len(row) else None)
        except ValueError as exc:
            warnings.append({"kind": "bad_row", "row": index, "item": first, "detail": str(exc)})
            continue

        key = (first, when)
        if key in seen:
            warnings.append({"kind": "duplicate_item_day", "row": index, "item": first})
            continue
        seen.add(key)

        net = 0
        if total_col is not None and total_col < len(row) and row[total_col] is not None:
            try:
                net = to_paise(row[total_col])
            except (ValueError, ArithmeticError):
                warnings.append({"kind": "bad_amount", "row": index, "item": first})

        rows.append(ParsedItemDay(item_name=first, report_date=when, qty=qty, net_paise=net))

    if not rows:
        raise UnreadableExport("The export contained no item-day rows.")

    return ItemDaysResult(
        rows=rows,
        warnings=warnings,
        period_start=period_start,
        period_end=period_end,
        restaurant=restaurant,
        reported_qty=reported_qty,
    )
