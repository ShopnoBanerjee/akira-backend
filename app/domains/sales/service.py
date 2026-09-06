"""Ingesting a Petpooja export.

The shape of this, and why:

    upload  ->  hash  ->  already seen?  ->  store  ->  row  ->  background parse

**Idempotent by content, not by filename.** `data_uploads.file_sha256` is
unique, so re-uploading the same export — which somebody will do, because
nothing about a spreadsheet tells you whether you already sent it — returns the
original upload instead of doubling six weeks of revenue. Filenames carry an
export timestamp and change every time; the bytes do not.

**Parsing never runs inside the request.** A six-week export is 452 bills and
growing; the same rule that keeps photo hashing out of the request path applies
here, and for the same reason — it is bracketed by a `job_runs` row so a failure
is something an admin can read rather than a lost upload.

**The uploader names the outlet.** The Orders Master Report has no outlet
column; its preamble says only "Akira", and both outlets would export files
that look identical. So the outlet is chosen explicitly at upload, scoped to
what the uploader can access.

**The source file is kept.** That is what `storage_path` is for: an adapter
version bump should be able to re-read the original rather than asking somebody
to export it again. It does mean the raw export — customer phone numbers and
all — sits in Storage, so the bucket is private and there is no browser read
path to it. Only the salted digest ever reaches a database column.
"""

import hashlib
import json
import uuid
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Any

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import record
from app.core.config import get_settings
from app.core.deps import CurrentUser
from app.core.enums import AuditAction, UploadStatus
from app.core.errors import ConflictError, ForbiddenError, NotFoundError, ValidationError
from app.core.settings_value import resolve
from app.domains.sales import (
    petpooja,
    petpooja_categories,
    petpooja_itemdays,
    petpooja_itemwise,
    petpooja_listing,
)
from app.integrations import storage

XLSX_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/octet-stream",
}
MAX_UPLOAD_BYTES = 25 * 1024 * 1024

SOURCE_ORDERS = "petpooja_orders"
SOURCE_LISTING = "petpooja_listing"
SOURCE_ITEMDAYS = "petpooja_itemdays"
SOURCE_CATEGORIES = "petpooja_categories"
SOURCE_ITEMWISE = "petpooja_itemwise"


@dataclass(frozen=True)
class ExportPreamble:
    """What can be learned about a file before committing to parsing it."""

    #: Which of the three reports this is.
    source: str
    #: The "Restaurant Name:" line, verbatim, or None if the file has none.
    restaurant: str | None


def inspect_export(data: bytes) -> ExportPreamble:
    """Which Petpooja report these bytes are, and which venue they claim.

    The uploader should not have to know Petpooja's report names — one upload
    button, and the file says what it is. The master's header has "Invoice
    No." and no "Items"; the listing has "Order No." plus "Items". Anything
    else is refused HERE, in the request, so whoever attached the wrong file
    hears about it now rather than finding a failed upload row tomorrow.

    The restaurant name is read in the same pass, and for the same reason. It
    sits in the preamble, above the header row, so the scan below has already
    walked past it by the time the report type is known — reading it costs
    nothing here and would cost a second 25MB workbook load anywhere else.
    """
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(f"Not a readable .xlsx file: {exc}") from exc
    seen: list[tuple[Any, ...]] = []
    source: str | None = None
    try:
        sheet = workbook[workbook.sheetnames[0]]
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= 30:  # headers sit in the first few rows of every export seen
                break
            seen.append(tuple(row))
            cells = {str(cell or "").strip() for cell in row}
            if petpooja.HEADER_MARKER in cells:
                source = SOURCE_ORDERS
                break
            if all(m in cells for m in petpooja_listing.LISTING_HEADER_MARKERS):
                source = SOURCE_LISTING
                break
            if all(m in cells for m in petpooja_categories.CATEGORIES_HEADER_MARKERS):
                source = SOURCE_CATEGORIES
                break
            # Item Wise has Category + Item + Qty. and no Date; Item Report:
            # Day Wise has Item + Date + Qty. and no Category. The Date column
            # is what tells them apart.
            if (
                all(m in cells for m in petpooja_itemwise.ITEMWISE_HEADER_MARKERS)
                and "Date" not in cells
            ):
                source = SOURCE_ITEMWISE
                break
            if all(m in cells for m in petpooja_itemdays.ITEMDAYS_HEADER_MARKERS):
                source = SOURCE_ITEMDAYS
                break
    finally:
        workbook.close()
    if source is None:
        raise ValidationError(
            "This is not a report this system can read. Supported, all from "
            "Petpooja: the Orders Master Report, the Order Listing report, the "
            "Item Report: Day Wise, the Item Wise: Sales Report, and the Sales "
            "Report: Category Wise."
        )
    return ExportPreamble(source=source, restaurant=petpooja.restaurant_in_preamble(seen))


def detect_source(data: bytes) -> str:
    """Which Petpooja report these bytes are. See `inspect_export`."""
    return inspect_export(data).source


async def check_restaurant(
    db: AsyncSession,
    *,
    outlet_id: uuid.UUID,
    found: str | None,
) -> None:
    """Refuse a file that names a restaurant this outlet does not expect.

    The threat is narrow and worth stating exactly, because a guard people
    misread is worse than none:

    - It CATCHES another venue's export — a second restaurant in the same
      Petpooja account, a file forwarded by the wrong person, an accountant's
      spreadsheet for somewhere else. That file would otherwise ingest
      silently and look, afterwards, like a slow month.
    - It CANNOT catch New Town's export filed against the other outlet. Both
      outlets sit under one Petpooja account and print the same name, which
      is the same reason the uploader has to pick the outlet by hand in the
      first place. Nothing in the file distinguishes them.

    Unarmed by default. An empty setting means every existing installation
    keeps working and nobody is locked out of uploading on the day this ships;
    the observed name is recorded regardless, which is what makes arming it a
    copy-paste rather than a guess.
    """
    expected = str(await resolve(db, "sales.petpooja_restaurant_name", outlet_id=outlet_id) or "")
    if not expected.strip():
        return
    if petpooja.normalise_restaurant(found) == petpooja.normalise_restaurant(expected):
        return
    wanted = expected.strip()
    if found:
        message = (
            f"This export is for {found!r}, but this outlet expects {wanted!r}. "
            "Nothing was ingested. If that is the same restaurant written "
            "differently, correct the 'Expected Petpooja restaurant name' "
            "setting; otherwise this is another venue's file."
        )
    else:
        message = (
            "This export carries no 'Restaurant Name:' line, so it cannot be "
            f"checked against the {wanted!r} this outlet expects. Nothing was "
            "ingested."
        )
    raise ValidationError(message, extra={"found": found, "expected": wanted})


def phone_hasher() -> "Any":
    """Salted SHA-256, per the spec's privacy note.

    The salt lives in the environment. Rotating it orphans every existing hash,
    which is the documented trade-off for it not being guessable from the
    digest alone.
    """
    salt = get_settings().PHONE_HASH_SALT

    def digest(digits: str) -> str:
        return hashlib.sha256(f"{salt}:{digits}".encode()).hexdigest()

    return digest


def storage_path_for(outlet_id: uuid.UUID, file_sha256: str) -> str:
    """Keyed by content hash, so the same file always lands in the same place
    and re-uploading overwrites rather than accumulating."""
    return f"{outlet_id}/{file_sha256}.xlsx"


async def create_upload(
    db: AsyncSession,
    user: CurrentUser,
    *,
    outlet_id: uuid.UUID,
    filename: str,
    content_type: str,
    data: bytes,
    **audit_ctx: Any,
) -> dict[str, Any]:
    """Accept a file, or hand back the upload that already holds it."""
    if not user.can_access_outlet(outlet_id):
        raise ForbiddenError("You do not have access to that outlet.")
    if content_type not in XLSX_TYPES and not filename.lower().endswith(".xlsx"):
        raise ValidationError(
            "Sales exports must be .xlsx files, as Petpooja produces them.",
            extra={"content_type": content_type},
        )
    if not data:
        raise ValidationError("That file is empty.")
    if len(data) > MAX_UPLOAD_BYTES:
        raise ValidationError(
            f"That file is over {MAX_UPLOAD_BYTES // (1024 * 1024)}MB.",
            extra={"bytes": len(data)},
        )

    file_sha256 = hashlib.sha256(data).hexdigest()
    preamble = inspect_export(data)
    source = preamble.source
    # Before the bytes reach Storage and before a row exists: a refused file
    # should leave no trace to clean up, and the person holding the wrong
    # spreadsheet should hear about it while they still remember sending it.
    await check_restaurant(db, outlet_id=outlet_id, found=preamble.restaurant)

    # Idempotency. Somebody will re-send the same export, because nothing about
    # a spreadsheet says whether it has been sent before.
    existing = (
        (
            await db.execute(
                text(
                    "select id, outlet_id, status, row_count, created_at"
                    " from data_uploads where file_sha256 = :h"
                ),
                {"h": file_sha256},
            )
        )
        .mappings()
        .first()
    )
    if existing is not None:
        if existing["outlet_id"] != outlet_id:
            # The same bytes cannot belong to two outlets. Refusing beats
            # silently filing a New Town export under another outlet.
            raise ConflictError(
                "This exact file has already been uploaded for a different outlet.",
                extra={"upload_id": str(existing["id"])},
            )
        return {
            **dict(existing),
            "id": str(existing["id"]),
            "already_ingested": True,
            "detail": "This file has been uploaded before; nothing was ingested twice.",
        }

    path = storage_path_for(outlet_id, file_sha256)
    await storage.upload_bytes(
        path,
        data,
        bucket=storage.SALES_UPLOAD_BUCKET,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    upload_id = (
        await db.execute(
            text(
                """
                insert into data_uploads
                    (outlet_id, uploaded_by, source, original_filename,
                     storage_path, file_sha256, status, restaurant_name)
                values (:outlet_id, :uploaded_by, :source, :filename,
                        :path, :sha, 'received', :restaurant)
                returning id
                """
            ),
            {
                "outlet_id": outlet_id,
                "uploaded_by": user.profile_id,
                "source": source,
                "filename": filename[:250],
                "path": path,
                "sha": file_sha256,
                "restaurant": preamble.restaurant,
            },
        )
    ).scalar_one()

    await record(
        db,
        actor_profile_id=user.profile_id,
        outlet_id=outlet_id,
        entity_table="data_uploads",
        entity_id=upload_id,
        action=AuditAction.CREATE,
        after={
            "filename": filename,
            "bytes": len(data),
            "sha256": file_sha256,
            "restaurant_name": preamble.restaurant,
        },
        **audit_ctx,
    )
    await db.commit()
    return {
        "id": str(upload_id),
        "outlet_id": outlet_id,
        "source": source,
        "status": UploadStatus.RECEIVED.value,
        "row_count": None,
        "already_ingested": False,
        "detail": "Received. Parsing runs in the background.",
    }


async def parse_upload(db: AsyncSession, upload_id: uuid.UUID) -> dict[str, Any]:
    """Read the stored file and write its bills. Idempotent per upload.

    Runs in a background task. Anything it raises is caught by `run_job` and
    written to `job_runs`; anything it can explain is written to the upload's
    own `error_detail`, which is what the person who sent the file will read.
    """
    upload = (
        (
            await db.execute(
                text(
                    "select id, outlet_id, storage_path, original_filename, source, status"
                    " from data_uploads where id = :id"
                ),
                {"id": upload_id},
            )
        )
        .mappings()
        .first()
    )
    if upload is None:
        raise NotFoundError("That upload does not exist.")

    # error_detail is cleared with the same statement. Leaving the previous
    # failure's message beside status='parsing' makes a re-parse in flight
    # read as a fresh failure, which is the row lying about itself.
    await db.execute(
        text("update data_uploads set status = 'parsing', error_detail = null where id = :id"),
        {"id": upload_id},
    )
    await db.commit()

    try:
        data = await storage.download_object(
            upload["storage_path"], bucket=storage.SALES_UPLOAD_BUCKET
        )
        if upload["source"] == SOURCE_LISTING:
            return await _parse_listing(db, upload, data)
        if upload["source"] == SOURCE_ITEMDAYS:
            return await _parse_itemdays(db, upload, data)
        if upload["source"] == SOURCE_CATEGORIES:
            return await _parse_categories(db, upload, data)
        if upload["source"] == SOURCE_ITEMWISE:
            return await _parse_itemwise(db, upload, data)
        result = petpooja.parse_orders(data, hash_phone=phone_hasher())
        # A second time, at the last point before anything is written. The
        # request-time check cannot cover a re-parse: the file was accepted
        # under whatever the setting said then, and re-reading it today must
        # obey what it says today.
        await check_restaurant(db, outlet_id=upload["outlet_id"], found=result.restaurant)
    except Exception as exc:
        await db.rollback()
        await db.execute(
            text(
                "update data_uploads set status = 'failed', error_detail = :err,"
                " parsed_at = now() where id = :id"
            ),
            {"id": upload_id, "err": f"{type(exc).__name__}: {exc}"[:2000]},
        )
        await db.commit()
        raise

    written = await _write_orders(db, upload["outlet_id"], upload_id, result)

    import json

    await db.execute(
        text(
            """
            update data_uploads
               set status = 'parsed',
                   row_count = :rows,
                   period_start = :start,
                   period_end = :end,
                   warnings = cast(:warnings as jsonb),
                   adapter_version = :adapter,
                   parsed_net_paise = :net,
                   reported_net_paise = :reported,
                   restaurant_name = :restaurant,
                   parsed_at = now(),
                   error_detail = null
             where id = :id
            """
        ),
        {
            "id": upload_id,
            "rows": len(result.orders),
            "start": result.period_start,
            "end": result.period_end,
            "warnings": json.dumps(result.warnings, default=str),
            "adapter": result.adapter_version,
            "reported": result.reported_net_paise,
            "net": result.net_paise,
            "restaurant": result.restaurant,
        },
    )
    await db.commit()

    return {
        "upload_id": str(upload_id),
        "orders_parsed": len(result.orders),
        "orders_written": written["inserted"],
        "orders_updated": written["updated"],
        "net_paise": result.net_paise,
        "period": [str(result.period_start), str(result.period_end)],
        "warnings": len(result.warnings),
        "adapter": result.adapter_version,
    }


#: One statement per chunk rather than one per bill. The first version of this
#: looped, and 452 bills took 75 seconds against Supabase — 452 round trips.
#: Six weeks of history is the small case; a year would have taken eleven
#: minutes and looked like a hang.
WRITE_CHUNK = 1000

_UPSERT_ORDERS = text(
    """
    insert into sales_orders
        (outlet_id, upload_id, external_bill_no, business_date, ordered_at,
         channel, covers, gross_paise, discount_paise, tax_paise, net_paise,
         payment_mode, table_no, customer_phone_hash)
    select :outlet_id, :upload_id, b.bill_no, b.business_date, b.ordered_at,
           cast(b.channel as sales_channel), b.covers, b.gross, b.discount,
           b.tax, b.net, b.payment, b.table_no, b.phone
      from unnest(
              cast(:bill_nos as text[]),
              cast(:business_dates as date[]),
              cast(:ordered_ats as timestamptz[]),
              cast(:channels as text[]),
              cast(:covers as integer[]),
              cast(:gross as bigint[]),
              cast(:discount as bigint[]),
              cast(:tax as bigint[]),
              cast(:net as bigint[]),
              cast(:payments as text[]),
              cast(:tables as text[]),
              cast(:phones as text[])
           ) as b(bill_no, business_date, ordered_at, channel, covers, gross,
                  discount, tax, net, payment, table_no, phone)
    on conflict (outlet_id, external_bill_no) do update
       set upload_id = excluded.upload_id,
           business_date = excluded.business_date,
           ordered_at = excluded.ordered_at,
           channel = excluded.channel,
           covers = excluded.covers,
           gross_paise = excluded.gross_paise,
           discount_paise = excluded.discount_paise,
           tax_paise = excluded.tax_paise,
           net_paise = excluded.net_paise,
           payment_mode = excluded.payment_mode,
           table_no = excluded.table_no,
           customer_phone_hash = excluded.customer_phone_hash
    returning (xmax = 0) as inserted
    """
)


async def _write_orders(
    db: AsyncSession,
    outlet_id: uuid.UUID,
    upload_id: uuid.UUID,
    result: petpooja.ParseResult,
) -> dict[str, int]:
    """Upsert every bill on (outlet_id, external_bill_no), set at a time.

    Petpooja exports overlap: the next one covers the same six weeks plus a few
    more days. A bill already present is updated in place rather than inserted
    again — the unique constraint would refuse it anyway, and skipping would
    leave a corrected bill showing its old total forever.

    `xmax = 0` is Postgres telling us which rows the upsert actually inserted
    rather than updated, so the count is the database's answer and not a guess
    made before the statement ran.
    """
    inserted = updated = 0
    orders = result.orders
    for offset in range(0, len(orders), WRITE_CHUNK):
        chunk = orders[offset : offset + WRITE_CHUNK]
        rows = (
            (
                await db.execute(
                    _UPSERT_ORDERS,
                    {
                        "outlet_id": outlet_id,
                        "upload_id": upload_id,
                        "bill_nos": [o.external_bill_no for o in chunk],
                        "business_dates": [o.business_date for o in chunk],
                        "ordered_ats": [o.ordered_at for o in chunk],
                        "channels": [o.channel.value if o.channel else None for o in chunk],
                        "covers": [o.covers for o in chunk],
                        "gross": [o.gross_paise for o in chunk],
                        "discount": [o.discount_paise for o in chunk],
                        "tax": [o.tax_paise for o in chunk],
                        "net": [o.net_paise for o in chunk],
                        "payments": [o.payment_mode for o in chunk],
                        "tables": [o.table_no for o in chunk],
                        "phones": [o.customer_phone_hash for o in chunk],
                    },
                )
            )
            .scalars()
            .all()
        )
        inserted += sum(1 for was_new in rows if was_new)
        updated += sum(1 for was_new in rows if not was_new)
    return {"inserted": inserted, "updated": updated}


async def _parse_listing(db: AsyncSession, upload: Any, data: bytes) -> dict[str, Any]:
    """The Order Listing path: item names onto bills the master already owns.

    Bills come from the Orders Master; this file only decorates them. An
    order the master has not seen gets a warning, not a row — inventing a
    bill from a names-only export would give it a net of nothing and a date
    of maybe, and both would look like data.
    """
    result = petpooja_listing.parse_listing(data)
    await check_restaurant(db, outlet_id=upload["outlet_id"], found=result.restaurant)
    written = await _write_items(db, upload["outlet_id"], upload["id"], result)

    import json

    warnings = list(result.warnings)
    if written["unmatched"]:
        warnings.append(
            {
                "kind": "unmatched_bills",
                "count": len(written["unmatched"]),
                "bill_nos": written["unmatched"][:50],
                "effect": (
                    "no items written for these — upload the Orders Master "
                    "Report covering this period first, then re-parse"
                ),
            }
        )

    period_start, period_end = result.period
    await db.execute(
        text(
            """
            update data_uploads
               set status = 'parsed',
                   row_count = :rows,
                   period_start = :start,
                   period_end = :end,
                   warnings = cast(:warnings as jsonb),
                   adapter_version = :adapter,
                   parsed_net_paise = :net,
                   restaurant_name = :restaurant,
                   parsed_at = now(),
                   error_detail = null
             where id = :id
            """
        ),
        {
            "id": upload["id"],
            "rows": len(result.orders),
            "start": period_start,
            "end": period_end,
            "warnings": json.dumps(warnings, default=str),
            "adapter": result.adapter_version,
            "restaurant": result.restaurant,
            # The sum of My Amount across parsed orders. Verified against the
            # real exports: this is the master's GROSS for the same bills
            # (net + discounts), kept so the uploads list shows the two
            # reports agreeing — the live file agreed to the paisa.
            "net": result.amount_paise,
        },
    )
    await db.commit()

    return {
        "upload_id": str(upload["id"]),
        "orders_parsed": len(result.orders),
        "orders_matched": written["matched"],
        "orders_unmatched": len(written["unmatched"]),
        "items_written": written["items"],
        "payment_split_rows": result.payment_split_rows,
        "warnings": len(warnings),
        "adapter": result.adapter_version,
    }


_INSERT_ITEMS = text(
    """
    insert into sales_order_items
        (order_id, outlet_id, business_date, item_name, sl_no, upload_id)
    select b.order_id, :outlet_id, b.business_date, b.item_name, b.sl_no, :upload_id
      from unnest(
              cast(:order_ids as uuid[]),
              cast(:business_dates as date[]),
              cast(:item_names as text[]),
              cast(:sl_nos as integer[])
           ) as b(order_id, business_date, item_name, sl_no)
    """
)


async def _write_items(
    db: AsyncSession,
    outlet_id: uuid.UUID,
    upload_id: uuid.UUID,
    result: "petpooja_listing.ListingResult",
) -> dict[str, Any]:
    """Replace each matched order's item set, one statement per step.

    Delete-and-insert rather than upsert: the listing is the whole truth
    about an order's items, so a re-parse must also remove a name that a
    corrected export no longer carries. Quantity and price stay null — the
    export does not say, and null is the honest spelling of that.

    The business_date on each item row comes from the MATCHED master bill,
    never from this file's own timestamp — one source of truth for dating.
    """
    bill_nos = [o.external_bill_no for o in result.orders]
    matched = (
        (
            await db.execute(
                text(
                    """
                    select id, external_bill_no, business_date from sales_orders
                     where outlet_id = :o and external_bill_no = any(:bills)
                    """
                ),
                {"o": outlet_id, "bills": bill_nos},
            )
        )
        .mappings()
        .all()
    )
    by_bill = {r["external_bill_no"]: r for r in matched}
    unmatched = [b for b in bill_nos if b not in by_bill]

    order_ids: list[uuid.UUID] = []
    dates: list[date] = []
    names: list[str] = []
    sl_nos: list[int] = []
    for order in result.orders:
        hit = by_bill.get(order.external_bill_no)
        if hit is None:
            continue
        for sl, name in enumerate(order.item_names, start=1):
            order_ids.append(hit["id"])
            dates.append(hit["business_date"])
            names.append(name)
            sl_nos.append(sl)

    if by_bill:
        await db.execute(
            text("delete from sales_order_items where order_id = any(:ids)"),
            {"ids": [r["id"] for r in matched]},
        )
    for offset in range(0, len(order_ids), WRITE_CHUNK):
        await db.execute(
            _INSERT_ITEMS,
            {
                "outlet_id": outlet_id,
                "upload_id": upload_id,
                "order_ids": order_ids[offset : offset + WRITE_CHUNK],
                "business_dates": dates[offset : offset + WRITE_CHUNK],
                "item_names": names[offset : offset + WRITE_CHUNK],
                "sl_nos": sl_nos[offset : offset + WRITE_CHUNK],
            },
        )
    return {"matched": len(by_bill), "unmatched": unmatched, "items": len(order_ids)}


async def _parse_itemdays(db: AsyncSession, upload: Any, data: bytes) -> dict[str, Any]:
    """The Item Report: Day Wise path — true units per menu item per day.

    Upserts on (outlet, report_date, item_name): exports overlap the same
    way the master's do, and a corrected day must overwrite, not duplicate.
    The report's own Total quantity is kept beside our sum on the upload row
    (via parsed figures in the result), so a disagreement is visible.
    """
    result = petpooja_itemdays.parse_itemdays(data)
    await check_restaurant(db, outlet_id=upload["outlet_id"], found=result.restaurant)
    written = await _write_item_days(db, upload["outlet_id"], upload["id"], result)

    import json

    await db.execute(
        text(
            """
            update data_uploads
               set status = 'parsed',
                   row_count = :rows,
                   period_start = :start,
                   period_end = :end,
                   warnings = cast(:warnings as jsonb),
                   adapter_version = :adapter,
                   parsed_net_paise = :net,
                   restaurant_name = :restaurant,
                   parsed_at = now(),
                   error_detail = null
             where id = :id
            """
        ),
        {
            "id": upload["id"],
            "rows": len(result.rows),
            "start": result.period_start,
            "end": result.period_end,
            "warnings": json.dumps(result.warnings, default=str),
            "adapter": result.adapter_version,
            "net": result.net_paise,
            "restaurant": result.restaurant,
        },
    )
    await db.commit()
    return {
        "upload_id": str(upload["id"]),
        "item_days_parsed": len(result.rows),
        "inserted": written["inserted"],
        "updated": written["updated"],
        "total_qty": result.total_qty,
        "reported_qty": result.reported_qty,
        "warnings": len(result.warnings),
        "adapter": result.adapter_version,
    }


_UPSERT_ITEM_DAYS = text(
    """
    insert into sales_item_days
        (outlet_id, report_date, item_name, qty, net_paise, upload_id)
    select :outlet_id, b.report_date, b.item_name, b.qty, b.net_paise, :upload_id
      from unnest(
              cast(:dates as date[]),
              cast(:names as text[]),
              cast(:qtys as numeric[]),
              cast(:nets as bigint[])
           ) as b(report_date, item_name, qty, net_paise)
    on conflict (outlet_id, report_date, item_name) do update
       set qty = excluded.qty,
           net_paise = excluded.net_paise,
           upload_id = excluded.upload_id
    returning (xmax = 0) as inserted
    """
)


async def _write_item_days(
    db: AsyncSession,
    outlet_id: uuid.UUID,
    upload_id: uuid.UUID,
    result: "petpooja_itemdays.ItemDaysResult",
) -> dict[str, int]:
    inserted = updated = 0
    rows = result.rows
    for offset in range(0, len(rows), WRITE_CHUNK):
        chunk = rows[offset : offset + WRITE_CHUNK]
        flags = (
            (
                await db.execute(
                    _UPSERT_ITEM_DAYS,
                    {
                        "outlet_id": outlet_id,
                        "upload_id": upload_id,
                        "dates": [r.report_date for r in chunk],
                        "names": [r.item_name for r in chunk],
                        "qtys": [r.qty for r in chunk],
                        "nets": [r.net_paise for r in chunk],
                    },
                )
            )
            .scalars()
            .all()
        )
        inserted += sum(1 for was_new in flags if was_new)
        updated += sum(1 for was_new in flags if not was_new)
    return {"inserted": inserted, "updated": updated}


_REPLACE_CATEGORY_PERIOD = text(
    """
    delete from sales_category_periods
     where outlet_id = :outlet_id and period_start = :start and period_end = :end
    """
)

_INSERT_CATEGORY_PERIODS = text(
    """
    insert into sales_category_periods
        (outlet_id, upload_id, period_start, period_end, category, orders, items,
         net_amount_paise, discount_paise, tax_paise, gross_paise, net_sales_paise,
         share_pct, is_charge)
    select :outlet_id, :upload_id, :start, :end,
           c, o, i, na, d, t, g, ns, sp, ch
      from unnest(
               cast(:categories as text[]), cast(:orders as int[]), cast(:items as int[]),
               cast(:net_amounts as bigint[]), cast(:discounts as bigint[]),
               cast(:taxes as bigint[]), cast(:grosses as bigint[]),
               cast(:net_sales as bigint[]), cast(:shares as float8[]),
               cast(:charges as boolean[])
           ) as u(c, o, i, na, d, t, g, ns, sp, ch)
    """
)


async def _parse_categories(db: AsyncSession, upload: Any, data: bytes) -> dict[str, Any]:
    """The Category Wise path: bills-per-category for one period (D29).

    A period's rows are REPLACED as a set, not upserted row by row: Petpooja
    recomputes the whole report for the range, so a re-export supersedes the
    old one — including a category that vanished from it.
    """
    result = petpooja_categories.parse_categories(data)
    await check_restaurant(db, outlet_id=upload["outlet_id"], found=result.restaurant)
    if result.period_start is None or result.period_end is None:
        raise petpooja.UnreadableExport(
            "The export has no readable 'Date:' line. A category report is a "
            "statement about a period; without one it cannot be stored."
        )

    params = {
        "outlet_id": upload["outlet_id"],
        "start": result.period_start,
        "end": result.period_end,
    }
    await db.execute(_REPLACE_CATEGORY_PERIOD, params)
    rows = result.rows
    await db.execute(
        _INSERT_CATEGORY_PERIODS,
        {
            **params,
            "upload_id": upload["id"],
            "categories": [r.category for r in rows],
            "orders": [r.orders for r in rows],
            "items": [r.items for r in rows],
            "net_amounts": [r.net_amount_paise for r in rows],
            "discounts": [r.discount_paise for r in rows],
            "taxes": [r.tax_paise for r in rows],
            "grosses": [r.gross_paise for r in rows],
            "net_sales": [r.net_sales_paise for r in rows],
            "shares": [r.share_pct for r in rows],
            "charges": [r.is_charge for r in rows],
        },
    )

    import json

    await db.execute(
        text(
            """
            update data_uploads
               set status = 'parsed',
                   row_count = :rows,
                   period_start = :start,
                   period_end = :end,
                   warnings = cast(:warnings as jsonb),
                   adapter_version = :adapter,
                   parsed_net_paise = :net,
                   reported_net_paise = :reported,
                   restaurant_name = :restaurant,
                   parsed_at = now(),
                   error_detail = null
             where id = :id
            """
        ),
        {
            "id": upload["id"],
            "rows": len(rows),
            "start": result.period_start,
            "end": result.period_end,
            "warnings": json.dumps(result.warnings, default=str),
            "adapter": result.adapter_version,
            "net": result.net_sales_paise,
            "reported": result.reported_net_sales_paise,
            "restaurant": result.restaurant,
        },
    )
    await db.commit()
    return {
        "upload_id": str(upload["id"]),
        "categories": len(rows),
        "period": [str(result.period_start), str(result.period_end)],
        "net_sales_paise": result.net_sales_paise,
        "warnings": len(result.warnings),
        "adapter": result.adapter_version,
    }


async def organisation_of_outlet(db: AsyncSession, outlet_id: uuid.UUID) -> uuid.UUID:
    """The tenant an outlet belongs to; the menu map and recipes are keyed by
    it (D33)."""
    org = await db.scalar(
        text("select organisation_id from outlets where id = :o"), {"o": outlet_id}
    )
    if org is None:
        raise NotFoundError("That outlet does not exist.")
    return uuid.UUID(str(org))


#: The organisation's own menu rows plus the starter kit's (D33).
_MENU_READ = "(m.organisation_id is null or m.organisation_id = cast(:org as uuid))"

_UPSERT_MENU_ITEMS = text(
    """
    insert into menu_items (organisation_id, name, category, petpooja_code, upload_id)
    select cast(:org as uuid), n, c, k, :upload_id
      from unnest(cast(:names as text[]), cast(:categories as text[]),
                  cast(:codes as text[])) as u(n, c, k)
    on conflict (coalesce(organisation_id, '00000000-0000-0000-0000-000000000000'),
                 lower(name)) do update set
        category      = excluded.category,
        petpooja_code = coalesce(excluded.petpooja_code, menu_items.petpooja_code),
        upload_id     = excluded.upload_id,
        updated_at    = now()
    returning (xmax = 0) as inserted
    """
)


async def _parse_itemwise(db: AsyncSession, upload: Any, data: bytes) -> dict[str, Any]:
    """The Item Wise path: the menu's own taxonomy (D29).

    Brand-level, keyed by Petpooja's printed name, upserted: an item that
    moves category is corrected, one that disappears from the menu keeps its
    last known category so old bills still join.
    """
    result = petpooja_itemwise.parse_itemwise(data)
    await check_restaurant(db, outlet_id=upload["outlet_id"], found=result.restaurant)
    organisation_id = await organisation_of_outlet(db, upload["outlet_id"])
    flags = (
        (
            await db.execute(
                _UPSERT_MENU_ITEMS,
                {
                    "org": organisation_id,
                    "upload_id": upload["id"],
                    "names": [i.item_name for i in result.items],
                    "categories": [i.category for i in result.items],
                    "codes": [i.petpooja_code for i in result.items],
                },
            )
        )
        .scalars()
        .all()
    )
    inserted = sum(1 for f in flags if f)

    import json

    await db.execute(
        text(
            """
            update data_uploads
               set status = 'parsed',
                   row_count = :rows,
                   period_start = :start,
                   period_end = :end,
                   warnings = cast(:warnings as jsonb),
                   adapter_version = :adapter,
                   parsed_net_paise = :net,
                   restaurant_name = :restaurant,
                   parsed_at = now(),
                   error_detail = null
             where id = :id
            """
        ),
        {
            "id": upload["id"],
            "rows": len(result.items),
            "start": result.period_start,
            "end": result.period_end,
            "warnings": json.dumps(result.warnings, default=str),
            "adapter": result.adapter_version,
            "net": sum(i.net_paise for i in result.items),
            "restaurant": result.restaurant,
        },
    )
    await db.commit()
    return {
        "upload_id": str(upload["id"]),
        "menu_items": len(result.items),
        "inserted": inserted,
        "updated": len(flags) - inserted,
        "categories": sorted(result.categories),
        "warnings": len(result.warnings),
        "adapter": result.adapter_version,
    }


async def menu_mix(
    db: AsyncSession,
    user: CurrentUser,
    *,
    outlet_id: uuid.UUID,
    date_from: date | None,
    date_to: date | None,
) -> dict[str, Any]:
    """Category attach for an outlet, two ways, each labelled with its source.

    **Reported** — Petpooja's Category Wise rows for one period: bills that
    carried the category, divided by the bills sales_orders holds for that
    period. The most recent period is used unless the caller names one that
    was uploaded. Petpooja counts on its own calendar dates and business_date
    rolls at 05:00, so the denominator can differ by a late-night bill or two
    at the edges; the response says which period it used.

    **Measured** — from the bills themselves: sales_order_items (names per
    bill, D21) joined to menu_items (name → category). Exact per bill, but
    only over bills whose items were uploaded. Names on bills that the menu
    map does not know are listed, because an unmapped item silently lowers
    every rate.
    """
    organisation_id = await organisation_of_outlet(db, outlet_id)
    if not user.can_access_outlet(outlet_id):
        raise ForbiddenError("You do not have access to that outlet.")

    period = (
        (
            await db.execute(
                text(
                    """
                    select period_start, period_end
                      from sales_category_periods
                     where outlet_id = :o
                       and (cast(:s as date) is null or period_start = :s)
                       and (cast(:e as date) is null or period_end = :e)
                     order by period_end desc, period_start asc
                     limit 1
                    """
                ),
                {"o": outlet_id, "s": date_from, "e": date_to},
            )
        )
        .mappings()
        .first()
    )

    reported: dict[str, Any] | None = None
    if period is not None:
        start, end = period["period_start"], period["period_end"]
        bills = int(
            await db.scalar(
                text(
                    "select count(*) from sales_orders"
                    " where outlet_id = :o and business_date between :s and :e"
                ),
                {"o": outlet_id, "s": start, "e": end},
            )
            or 0
        )
        rows = (
            await db.execute(
                text(
                    """
                    select category, orders, items, net_sales_paise, gross_paise,
                           cast(share_pct as float8) as share_pct, is_charge
                      from sales_category_periods
                     where outlet_id = :o and period_start = :s and period_end = :e
                     order by is_charge, orders desc, category
                    """
                ),
                {"o": outlet_id, "s": start, "e": end},
            )
        ).mappings()
        categories = []
        for r in rows:
            orders = int(r["orders"])
            categories.append(
                {
                    "category": r["category"],
                    "orders": orders,
                    "share_of_bills": (round(orders / bills, 4) if bills and orders else None),
                    "items": int(r["items"]),
                    "items_per_order": (round(int(r["items"]) / orders, 2) if orders else None),
                    "net_sales_paise": int(r["net_sales_paise"]),
                    "avg_spend_per_bill_paise": (
                        int(r["net_sales_paise"]) // orders if orders else None
                    ),
                    "share_of_net_pct": r["share_pct"],
                    "is_charge": bool(r["is_charge"]),
                }
            )
        reported = {
            "period_start": start,
            "period_end": end,
            "bills_in_period": bills,
            "categories": categories,
        }

    # Measured per bill, over whatever window the caller asked for (or the
    # reported period, or everything).
    m_start = date_from or (period["period_start"] if period else None)
    m_end = date_to or (period["period_end"] if period else None)
    measured_rows = (
        (
            await db.execute(
                text(
                    """
                with bills as (
                    select o.id
                      from sales_orders o
                     where o.outlet_id = :o
                       and (cast(:s as date) is null or o.business_date >= :s)
                       and (cast(:e as date) is null or o.business_date <= :e)
                       and exists (select 1 from sales_order_items i where i.order_id = o.id)
                ),
                cat as (
                    -- The printed name is the menu name, or an alias of one
                    -- (case-insensitive). Either way the category is the
                    -- menu item's.
                    select i.order_id, m.category
                      from sales_order_items i
                      join bills b on b.id = i.order_id
                      left join menu_items direct on direct.name = i.item_name
                      left join menu_item_aliases a on lower(a.alias) = lower(i.item_name)
                      join menu_items m on m.id = coalesce(direct.id, a.menu_item_id)
                )
                select category,
                       count(distinct order_id) as bills_with,
                       (select count(*) from bills) as bills_measured
                  from cat
                 group by category
                 order by bills_with desc, category
                """
                ),
                {"o": outlet_id, "s": m_start, "e": m_end},
            )
        )
        .mappings()
        .all()
    )
    bills_measured = int(measured_rows[0]["bills_measured"]) if measured_rows else 0
    unmapped = [
        r["item_name"]
        for r in (
            await db.execute(
                text(
                    f"""
                    select distinct i.item_name
                      from sales_order_items i
                      join sales_orders o on o.id = i.order_id
                     where o.outlet_id = :o
                       and not exists (
                           select 1 from menu_items m
                            where m.name = i.item_name and {_MENU_READ}
                       )
                       and not exists (
                           select 1 from menu_item_aliases m
                            where lower(m.alias) = lower(i.item_name) and {_MENU_READ}
                       )
                     order by 1
                    """
                ),
                {"o": outlet_id, "org": organisation_id},
            )
        ).mappings()
    ]
    measured = {
        "from": m_start,
        "to": m_end,
        "bills_measured": bills_measured,
        "categories": [
            {
                "category": r["category"],
                "bills_with": int(r["bills_with"]),
                "share_of_bills": round(int(r["bills_with"]) / bills_measured, 4)
                if bills_measured
                else None,
            }
            for r in measured_rows
        ],
        "unmapped_item_names": unmapped,
    }
    menu_known = int(
        await db.scalar(
            text(f"select count(*) from menu_items m where {_MENU_READ}"),
            {"org": organisation_id},
        )
        or 0
    )
    return {
        "outlet_id": str(outlet_id),
        "menu_items_known": menu_known,
        "reported": reported,
        "measured": measured,
    }


async def list_menu_items(
    db: AsyncSession, *, organisation_id: uuid.UUID | None
) -> list[dict[str, Any]]:
    """The menu map with its aliases, for the alias editor and the unmapped
    list's dropdown. Organisation-level, so no outlet filter."""
    rows = (
        await db.execute(
            text(
                f"""
                select m.id, m.name, m.category, m.petpooja_code,
                       coalesce(
                           json_agg(json_build_object('id', a.id, 'alias', a.alias)
                                    order by a.alias)
                           filter (where a.id is not null),
                           '[]'::json
                       ) as aliases
                  from menu_items m
                  left join menu_item_aliases a on a.menu_item_id = m.id
                 where {_MENU_READ}
                 group by m.id
                 order by m.category, m.name
                """
            ),
            {"org": organisation_id},
        )
    ).mappings()
    out = []
    for r in rows:
        d = dict(r)
        raw = d["aliases"]
        d["aliases"] = json.loads(raw) if isinstance(raw, str) else raw
        out.append(d)
    return out


async def add_menu_alias(
    db: AsyncSession,
    user: CurrentUser,
    *,
    alias: str,
    menu_item_name: str,
    ip: str | None = None,
    user_agent: str | None = None,
) -> dict[str, Any]:
    """Remember that a bill's spelling means this menu item.

    Refused when the alias IS a menu name (nothing to map) or is already
    mapped elsewhere (one spelling, one item). Case-insensitive throughout,
    because Petpooja's own casing is not consistent between reports.
    """
    alias = " ".join(alias.split())
    if not alias:
        raise ValidationError("The alias is empty.")
    item = (
        (
            await db.execute(
                text(
                    "select m.id, m.name from menu_items m"
                    f" where lower(m.name) = lower(:n) and {_MENU_READ}"
                ),
                {"n": menu_item_name, "org": user.organisation_id},
            )
        )
        .mappings()
        .first()
    )
    if item is None:
        raise NotFoundError(
            f"{menu_item_name!r} is not a menu item. Upload an Item Wise report that names it."
        )
    if await db.scalar(
        text(f"select 1 from menu_items m where lower(m.name) = lower(:a) and {_MENU_READ}"),
        {"a": alias, "org": user.organisation_id},
    ):
        raise ConflictError(f"{alias!r} is already a menu item name; it needs no alias.")
    existing = (
        (
            await db.execute(
                text(
                    """
                    select a.alias, m.name
                      from menu_item_aliases a join menu_items m on m.id = a.menu_item_id
                     where lower(a.alias) = lower(:a)
                       and (a.organisation_id is null
                            or a.organisation_id = cast(:org as uuid))
                    """
                ),
                {"a": alias, "org": user.organisation_id},
            )
        )
        .mappings()
        .first()
    )
    if existing is not None:
        raise ConflictError(
            f"{existing['alias']!r} already maps to {existing['name']!r}. "
            "Remove that alias first if it is wrong.",
            extra={"maps_to": existing["name"]},
        )
    alias_id = await db.scalar(
        text(
            """
            insert into menu_item_aliases (organisation_id, menu_item_id, alias, created_by)
            values (:org, :item_id, :alias, :by) returning id
            """
        ),
        {"item_id": item["id"], "alias": alias, "by": user.profile_id, "org": user.organisation_id},
    )
    await record(
        db,
        actor_profile_id=user.profile_id,
        entity_table="menu_item_aliases",
        entity_id=alias_id,
        action=AuditAction.CREATE,
        after={"alias": alias, "menu_item": item["name"]},
        ip=ip,
        user_agent=user_agent,
    )
    await db.commit()
    return {"id": str(alias_id), "alias": alias, "menu_item": item["name"]}


async def delete_menu_alias(
    db: AsyncSession,
    user: CurrentUser,
    alias_id: uuid.UUID,
    *,
    ip: str | None = None,
    user_agent: str | None = None,
) -> None:
    before = (
        (
            await db.execute(
                text(
                    """
                    select a.alias, m.name
                      from menu_item_aliases a join menu_items m on m.id = a.menu_item_id
                     where a.id = :id and a.organisation_id = cast(:org as uuid)
                    """
                ),
                {"id": alias_id, "org": user.organisation_id},
            )
        )
        .mappings()
        .first()
    )
    if before is None:
        raise NotFoundError("That alias does not exist.")
    await db.execute(text("delete from menu_item_aliases where id = :id"), {"id": alias_id})
    await record(
        db,
        actor_profile_id=user.profile_id,
        entity_table="menu_item_aliases",
        entity_id=alias_id,
        action=AuditAction.DELETE,
        before={"alias": before["alias"], "menu_item": before["name"]},
        ip=ip,
        user_agent=user_agent,
    )
    await db.commit()


async def background_parse(upload_id: uuid.UUID, outlet_id: uuid.UUID) -> None:
    """What the upload endpoint hands to a BackgroundTask."""

    async def body(db: AsyncSession) -> dict[str, Any]:
        return await parse_upload(db, upload_id)

    from app.jobs.runner import run_job

    await run_job("sales_ingest", body, outlet_id=outlet_id)


# ---------------------------------------------------------------------------
# Reading it back
# ---------------------------------------------------------------------------


async def list_uploads(
    db: AsyncSession, user: CurrentUser, *, outlet_id: uuid.UUID | None, limit: int
) -> list[dict[str, Any]]:
    clauses: list[str] = []
    params: dict[str, Any] = {"limit": limit}
    if outlet_id is not None:
        if not user.can_access_outlet(outlet_id):
            raise ForbiddenError("You do not have access to that outlet.")
        clauses.append("u.outlet_id = :outlet_id")
        params["outlet_id"] = outlet_id
    elif not user.is_platform_admin:
        if not user.outlet_ids:
            return []
        clauses.append("u.outlet_id = any(:ids)")
        params["ids"] = sorted(user.outlet_ids)
    where = ("where " + " and ".join(clauses)) if clauses else ""

    import json

    rows = (
        (
            await db.execute(
                text(
                    f"""
                    select u.id, u.outlet_id, o.code as outlet_code, u.source,
                           u.original_filename, u.file_sha256, u.status,
                           u.row_count, u.period_start, u.period_end,
                           u.warnings, u.error_detail, u.adapter_version,
                           u.parsed_net_paise, u.restaurant_name,
                           u.created_at, u.parsed_at,
                           p.full_name as uploaded_by_name
                      from data_uploads u
                      join outlets o on o.id = u.outlet_id
                      left join profiles p on p.id = u.uploaded_by
                     {where}
                     order by u.created_at desc
                     limit :limit
                    """
                ),
                params,
            )
        )
        .mappings()
        .all()
    )
    out = []
    for row in rows:
        data = dict(row)
        raw = data.get("warnings")
        data["warnings"] = json.loads(raw) if isinstance(raw, str) else (raw or [])
        out.append(data)
    return out


async def list_orders(
    db: AsyncSession,
    user: CurrentUser,
    *,
    outlet_id: uuid.UUID | None,
    date_from: date | None,
    date_to: date | None,
    limit: int,
) -> list[dict[str, Any]]:
    """The raw table view. Newest trading day first, newest bill first."""
    clauses: list[str] = []
    params: dict[str, Any] = {"limit": limit}
    if outlet_id is not None:
        if not user.can_access_outlet(outlet_id):
            raise ForbiddenError("You do not have access to that outlet.")
        clauses.append("s.outlet_id = :outlet_id")
        params["outlet_id"] = outlet_id
    elif not user.is_platform_admin:
        if not user.outlet_ids:
            return []
        clauses.append("s.outlet_id = any(:ids)")
        params["ids"] = sorted(user.outlet_ids)
    if date_from is not None:
        clauses.append("s.business_date >= :date_from")
        params["date_from"] = date_from
    if date_to is not None:
        clauses.append("s.business_date <= :date_to")
        params["date_to"] = date_to
    where = ("where " + " and ".join(clauses)) if clauses else ""

    rows = (
        (
            await db.execute(
                text(
                    f"""
                    select s.id, s.outlet_id, o.code as outlet_code,
                           s.external_bill_no, s.business_date, s.ordered_at,
                           s.channel, s.covers, s.gross_paise, s.discount_paise,
                           s.tax_paise, s.net_paise, s.payment_mode, s.table_no,
                           s.customer_phone_hash is not null as has_phone,
                           coalesce(it.items, cast('{{}}' as text[])) as items
                      from sales_orders s
                      join outlets o on o.id = s.outlet_id
                      left join lateral (
                           select array_agg(i.item_name order by i.sl_no) as items
                             from sales_order_items i
                            where i.order_id = s.id
                      ) it on true
                     {where}
                     order by s.business_date desc, s.ordered_at desc
                     limit :limit
                    """
                ),
                params,
            )
        )
        .mappings()
        .all()
    )
    return [dict(r) for r in rows]


async def daily_totals(
    db: AsyncSession,
    user: CurrentUser,
    *,
    outlet_id: uuid.UUID,
    date_from: date | None,
    date_to: date | None,
) -> list[dict[str, Any]]:
    """Net sales per trading day. Grouped by `business_date`, never by
    `ordered_at::date` — that would split every night at midnight."""
    if not user.can_access_outlet(outlet_id):
        raise ForbiddenError("You do not have access to that outlet.")
    rows = (
        (
            await db.execute(
                text(
                    """
                    select business_date,
                           count(*) as bills,
                           coalesce(sum(net_paise), 0) as net_paise,
                           coalesce(sum(covers), 0) as covers
                      from sales_orders
                     where outlet_id = :outlet_id
                       and (cast(:date_from as date) is null
                            or business_date >= cast(:date_from as date))
                       and (cast(:date_to as date) is null
                            or business_date <= cast(:date_to as date))
                     group by business_date
                     order by business_date desc
                    """
                ),
                {"outlet_id": outlet_id, "date_from": date_from, "date_to": date_to},
            )
        )
        .mappings()
        .all()
    )
    return [dict(r) for r in rows]


async def item_summary(
    db: AsyncSession,
    user: CurrentUser,
    *,
    outlet_id: uuid.UUID,
    date_from: date | None,
    date_to: date | None,
) -> list[dict[str, Any]]:
    """How often each item appears on a bill. Not how many were sold.

    The Order Listing carries names without quantities, so the honest unit is
    "bills carrying this item" — a bill with two Shoyu Ramen counts once. The
    column is named `bills` for exactly that reason; the day a true line-item
    export exists, a `qty_sold` column can appear beside it.
    """
    if not user.can_access_outlet(outlet_id):
        raise ForbiddenError("You do not have access to that outlet.")
    rows = (
        (
            await db.execute(
                text(
                    """
                    select item_name,
                           count(distinct order_id) as bills,
                           min(business_date) as first_date,
                           max(business_date) as last_date
                      from sales_order_items
                     where outlet_id = :outlet_id
                       and (cast(:date_from as date) is null
                            or business_date >= cast(:date_from as date))
                       and (cast(:date_to as date) is null
                            or business_date <= cast(:date_to as date))
                     group by item_name
                     order by bills desc, item_name
                    """
                ),
                {"outlet_id": outlet_id, "date_from": date_from, "date_to": date_to},
            )
        )
        .mappings()
        .all()
    )
    return [dict(r) for r in rows]
