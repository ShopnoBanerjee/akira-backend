"""Reading a Petpooja "Sales Report: Category Wise" export.

The one report that says, for a whole period, how many BILLS carried each
menu category. That is the attach rate — the share of tickets with a drink,
with a dessert — which no bill-level export gives directly and which the
owner named as the KPI they actually want.

**What this export actually looks like**, established by reading AKIRA's own
file (17 Jul to 5 Sep 2026):

    r0-r2   preamble — "Date:", "Name:", "Restaurant Name:"
    r3-r4   blank
    r5      the header row, 9 columns
    r6-r9   Total / Min. / Max. / Avg. — summary rows that are NOT categories
    r10+    one row per category, then three pseudo-categories

Three things the file taught, each with a guard below:

- **"No. of Orders" is per category, and the Total row SUMS it.** 413 bills
  had Ramen, 418 had Gyoza; the Total says 1,915, which is the sum of those
  counts, not a count of bills. A bill with three categories counts three
  times there. So the denominator for "share of bills" is NOT in this file —
  the caller takes it from `sales_orders` for the same period.
- **The last rows are charges, not categories.** "Container Charge",
  "Round Off" and "Waived Off" carry zero orders and zero items but a Total
  Sales figure. They are kept (they explain the gross) and flagged
  `is_charge`, so nobody computes an attach rate for rounding.
- **Two nets.** "Net Amount" is before discount; "Net Sales (N.A - T.D)" is
  after. Both are stored, named as the file names them.

Pure. Bytes in, dataclasses out.
"""

from dataclasses import dataclass, field
from datetime import date
from io import BytesIO
from typing import Any

from app.domains.sales.petpooja import (
    SUMMARY_LABELS,
    UnreadableExport,
    parse_period,
    restaurant_in_preamble,
    to_paise,
)

CATEGORIES_ADAPTER_VERSION = "petpooja.categories.v1"

#: Identify the header row. "No. of Orders" is what no other report has.
CATEGORIES_HEADER_MARKERS = ("Category", "No. of Orders", "Total Items Ordered")

COLUMNS = {
    "category": "Category",
    "orders": "No. of Orders",
    "items": "Total Items Ordered",
    "net_amount": "Net Amount (₹)",
    "discount": "Total Discount (₹)",
    "tax": "Total Tax (₹)",
    "gross": "Total Sales (₹)",
    "net_sales": "Net Sales (₹)(N.A - T.D)",
    "share": "Percentage (%)",
}
REQUIRED = ("category", "orders", "items")


@dataclass(frozen=True)
class ParsedCategory:
    category: str
    #: Bills that carried at least one item of this category.
    orders: int
    #: Units of this category sold across those bills.
    items: int
    net_amount_paise: int
    discount_paise: int
    tax_paise: int
    gross_paise: int
    net_sales_paise: int
    share_pct: float | None
    #: Container Charge / Round Off / Waived Off: money, not menu.
    is_charge: bool


@dataclass(frozen=True)
class CategoriesResult:
    rows: list[ParsedCategory]
    warnings: list[dict[str, Any]] = field(default_factory=list)
    period_start: date | None = None
    period_end: date | None = None
    restaurant: str | None = None
    adapter_version: str = CATEGORIES_ADAPTER_VERSION
    #: The file's own Total row: items and net sales. Orders is deliberately
    #: not kept from there — it is a sum of per-category counts, not bills.
    reported_items: int | None = None
    reported_net_sales_paise: int | None = None

    @property
    def total_items(self) -> int:
        return sum(r.items for r in self.rows)

    @property
    def net_sales_paise(self) -> int:
        return sum(r.net_sales_paise for r in self.rows)


def _int(value: object) -> int:
    if value is None or value == "":
        return 0
    return int(float(str(value)))


def _float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value))
    except ValueError:
        return None


def parse_categories(data: bytes) -> CategoriesResult:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise UnreadableExport(f"Not a readable .xlsx file: {exc}") from exc
    try:
        sheet = workbook[workbook.sheetnames[0]]
        raw = [tuple(r) for r in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    warnings: list[dict[str, Any]] = []
    period_start = period_end = None
    header: list[str] | None = None
    header_at = -1
    for index, row in enumerate(raw):
        label = str(row[0] or "").strip().rstrip(":").lower() if row else ""
        if label == "date" and len(row) > 1:
            period_start, period_end = parse_period(row[1])
        cells = {str(cell or "").strip() for cell in row}
        if all(marker in cells for marker in CATEGORIES_HEADER_MARKERS):
            header = [str(cell or "").strip() for cell in row]
            header_at = index
            break
    if header is None:
        raise UnreadableExport(
            "No header row containing 'Category', 'No. of Orders' and 'Total Items "
            "Ordered'. This does not look like a Petpooja Sales Report: Category Wise."
        )
    restaurant = restaurant_in_preamble(raw[:header_at])

    at: dict[str, int] = {}
    for key, name in COLUMNS.items():
        if name in header:
            at[key] = header.index(name)
        elif key in REQUIRED:
            raise UnreadableExport(f"The export has no {name!r} column.")
        else:
            warnings.append({"kind": "missing_column", "column": name, "effect": f"{key} not set"})

    def cell(row: tuple[Any, ...], key: str) -> Any:
        pos = at.get(key)
        return row[pos] if pos is not None and pos < len(row) else None

    rows: list[ParsedCategory] = []
    seen: set[str] = set()
    reported_items: int | None = None
    reported_net: int | None = None
    for index, row in enumerate(raw[header_at + 1 :], start=header_at + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        first = str(row[0] or "").strip()
        if first.lower() in SUMMARY_LABELS:
            if first.lower() == "total" and reported_items is None:
                try:
                    reported_items = _int(cell(row, "items"))
                    reported_net = to_paise(cell(row, "net_sales"))
                except (ValueError, ArithmeticError):
                    warnings.append({"kind": "bad_total_row", "row": index})
            continue
        if not first:
            continue
        key = first.casefold()
        if key in seen:
            warnings.append({"kind": "duplicate_category", "row": index, "category": first})
            continue
        seen.add(key)
        try:
            orders = _int(cell(row, "orders"))
            items = _int(cell(row, "items"))
            rows.append(
                ParsedCategory(
                    category=first,
                    orders=orders,
                    items=items,
                    net_amount_paise=to_paise(cell(row, "net_amount")),
                    discount_paise=to_paise(cell(row, "discount")),
                    tax_paise=to_paise(cell(row, "tax")),
                    gross_paise=to_paise(cell(row, "gross")),
                    net_sales_paise=to_paise(cell(row, "net_sales")),
                    share_pct=_float(cell(row, "share")),
                    is_charge=orders == 0 and items == 0,
                )
            )
        except (ValueError, ArithmeticError) as exc:
            warnings.append(
                {"kind": "bad_row", "row": index, "category": first, "detail": str(exc)}
            )

    if not rows:
        raise UnreadableExport("The export contained no category rows.")
    if reported_items is not None and reported_items != sum(r.items for r in rows):
        warnings.append(
            {
                "kind": "items_total_mismatch",
                "reported": reported_items,
                "parsed": sum(r.items for r in rows),
            }
        )
    return CategoriesResult(
        rows=rows,
        warnings=warnings,
        period_start=period_start,
        period_end=period_end,
        restaurant=restaurant,
        reported_items=reported_items,
        reported_net_sales_paise=reported_net,
    )
