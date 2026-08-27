"""Reading a Petpooja Orders Master Report.

A versioned adapter, not a general spreadsheet reader. The spec's risk table
says a Petpooja format change should be a new adapter rather than a rewrite,
so everything specific to this export's shape lives here and behind
`ORDERS_ADAPTER_VERSION`.

Pure. Bytes in, dataclasses out, no database and no network — which is what
lets the reconciliation be asserted against the real 452-bill export in a test
rather than argued about.

**What this export actually looks like**, established by reading AKIRA's own
file rather than from documentation:

    r0-r2   preamble — "Date:", "Name:", "Restaurant Name:"
    r3-r4   blank
    r5      the header row, 43 columns
    r6-r9   Total / Min. / Max. / Avg. — summary rows that are NOT bills
    r10+    one row per bill, newest first

The four summary rows are the trap. A parser that starts at the row after the
header ingests the file's own totals as a ₹4.9 lakh order, and the number is
plausible enough to survive review.

Three more things this file taught, all of which have a guard below:

- **Money arrives as float rupees.** `int(81.6 * 100)` is 8159, not 8160, and
  seven cells in the real export lose a paisa that way. Everything goes through
  `Decimal(str(v))`, which is exact by construction rather than correct by luck.
- **Bills after midnight belong to the previous trading day.** 21 of the 452
  are struck before 05:00. The business date comes from `business_date()`, never
  from the calendar date in the cell.
- **The file carries customer phone numbers and names** — 141 and 142 of them.
  The raw number never leaves this module: a hasher is injected and only the
  digest is returned.
"""

import re
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from app.core.business_date import OUTLET_TZ
from app.core.business_date import business_date as to_business_date
from app.core.enums import SalesChannel

#: Bumped when this adapter's understanding of the export changes. Stored on
#: the upload so a re-parse under a new adapter is distinguishable from the
#: original read.
ORDERS_ADAPTER_VERSION = "petpooja.orders.v1"

#: The column that identifies the header row. Searched for rather than assumed
#: at a fixed offset, because the preamble has already varied in length between
#: report types.
HEADER_MARKER = "Invoice No."

#: First-column values that mark a summary row rather than a bill.
SUMMARY_LABELS = frozenset({"total", "min.", "max.", "avg.", "sub total", "grand total"})

#: Petpooja's order types, as they appear in the file.
CHANNELS: dict[str, SalesChannel] = {
    "dine in": SalesChannel.DINE_IN,
    "dinein": SalesChannel.DINE_IN,
    "pick up": SalesChannel.PICKUP,
    "pickup": SalesChannel.PICKUP,
    "takeaway": SalesChannel.PICKUP,
    "delivery": SalesChannel.DELIVERY,
    "delivery(parcel)": SalesChannel.DELIVERY,
    "parcel": SalesChannel.DELIVERY,
}

#: Only completed bills count as sales. Anything else is reported and skipped
#: rather than quietly folded into the totals.
SUCCESS_STATUSES = frozenset({"success", "settled", "complete", "completed", "paid"})

COLUMNS = {
    "bill_no": "Invoice No.",
    "ordered_at": "Date",
    "payment": "Payment Type",
    "order_type": "Order Type",
    "status": "Status",
    "area": "Area",
    "covers": "Persons",
    "phone": "Phone",
    "gross": "My Amount (₹)",
    "discount": "Discount (₹)",
    "net": "Net Sales (₹)(M.A - D)",
    "tax": "Total Tax (₹)",
}
#: Without these the file is not an Orders Master Report and parsing stops.
REQUIRED = ("bill_no", "ordered_at", "net")


class UnreadableExport(ValueError):
    """The file is not a Petpooja Orders Master Report.

    Raised rather than returning an empty result: zero orders and "this is the
    wrong file" look identical in a database and must not.
    """


@dataclass(frozen=True)
class ParsedOrder:
    external_bill_no: str
    ordered_at: datetime
    business_date: date
    channel: SalesChannel | None
    covers: int | None
    gross_paise: int
    discount_paise: int
    tax_paise: int
    net_paise: int
    payment_mode: str | None
    table_no: str | None
    #: Already a digest. The raw number never leaves the parse loop.
    customer_phone_hash: str | None


@dataclass(frozen=True)
class ParseResult:
    orders: list[ParsedOrder]
    #: Everything unrecognised or skipped, shown to whoever uploaded the file.
    #: A parser that silently drops rows is a parser nobody can trust.
    warnings: list[dict[str, Any]] = field(default_factory=list)
    period_start: date | None = None
    period_end: date | None = None
    restaurant: str | None = None
    adapter_version: str = ORDERS_ADAPTER_VERSION
    #: The Net Sales figure from the export's own Total row, in paise. What
    #: the file CLAIMS it adds up to, kept apart from what we derived so a
    #: disagreement is a visible column pair, not a recomputation by hand.
    reported_net_paise: int | None = None

    @property
    def net_paise(self) -> int:
        return sum(o.net_paise for o in self.orders)


def to_paise(value: object) -> int:
    """Rupees to integer paise, exactly.

    `Decimal(str(v))` rather than `v * 100`: the float nearest to 81.6 is
    slightly below it, so multiplying and truncating loses a paisa. Seven cells
    in AKIRA's own export do exactly that.
    """
    if value is None or value == "":
        return 0
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("₹", "").strip()
        if not cleaned:
            return 0
        value = cleaned
    return int((Decimal(str(value)) * 100).to_integral_value())


def parse_timestamp(value: object) -> datetime:
    """The cell as an outlet-local aware datetime.

    openpyxl hands back either a datetime or a string depending on how the cell
    was written, and Petpooja writes both across report types.
    """
    if isinstance(value, datetime):
        stamped = value
    else:
        text = str(value).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d-%m-%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
            try:
                stamped = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        else:
            raise ValueError(f"unrecognised timestamp: {value!r}")
    # Petpooja exports wall-clock time at the restaurant, with no zone on it.
    return stamped if stamped.tzinfo else stamped.replace(tzinfo=OUTLET_TZ)


def parse_period(value: object) -> tuple[date | None, date | None]:
    """ "2026-07-17 to 2026-08-25" from the preamble."""
    found = re.findall(r"\d{4}-\d{2}-\d{2}", str(value or ""))
    if not found:
        return None, None
    first = date.fromisoformat(found[0])
    return first, date.fromisoformat(found[-1]) if len(found) > 1 else first


def _cell(row: tuple[Any, ...], at: dict[str, int], key: str) -> Any:
    """One column of one row by name, tolerating a short row.

    A module-level function rather than a closure over the loop variable: the
    closure form works only because it is called immediately, and the day
    somebody defers one it would silently read the last row of the file.
    """
    position = at.get(key)
    if position is None or position >= len(row):
        return None
    return row[position]


def _is_summary(first_cell: object) -> bool:
    return str(first_cell or "").strip().lower() in SUMMARY_LABELS


def _covers(value: object) -> int | None:
    try:
        n = int(float(str(value)))
    except (TypeError, ValueError):
        return None
    return n if n >= 0 else None


def parse_orders(data: bytes, *, hash_phone: Callable[[str], str] | None = None) -> ParseResult:
    """Read an Orders Master Report.

    `hash_phone` is injected so this module never has to know the salt and the
    raw number never exists outside the loop below. Without one, phones are
    dropped rather than returned in the clear.
    """
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise UnreadableExport(f"Not a readable .xlsx file: {exc}") from exc

    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    warnings: list[dict[str, Any]] = []
    period_start = period_end = None
    restaurant = None
    header: list[str] | None = None
    header_at = -1

    for index, row in enumerate(rows):
        label = str(row[0] or "").strip().rstrip(":").lower() if row else ""
        if label == "date" and len(row) > 1:
            period_start, period_end = parse_period(row[1])
        elif label == "restaurant name" and len(row) > 1:
            restaurant = str(row[1] or "").strip() or None
        if any(str(cell or "").strip() == HEADER_MARKER for cell in row):
            header = [str(cell or "").strip() for cell in row]
            header_at = index
            break

    if header is None:
        raise UnreadableExport(
            f"No header row containing {HEADER_MARKER!r}. "
            "This does not look like a Petpooja Orders Master Report."
        )

    at: dict[str, int] = {}
    for key, name in COLUMNS.items():
        if name in header:
            at[key] = header.index(name)
        elif key in REQUIRED:
            raise UnreadableExport(f"The export has no {name!r} column.")
        else:
            warnings.append({"kind": "missing_column", "column": name, "effect": f"{key} not set"})

    orders: list[ParsedOrder] = []
    seen: set[str] = set()
    reported_net: int | None = None

    for index, row in enumerate(rows[header_at + 1 :], start=header_at + 2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        if _is_summary(row[0]):
            # The file's own Total/Min/Max/Avg. Ingesting these as bills is the
            # single most damaging mistake this parser could make — but the
            # Total row's Net Sales is worth keeping: it is the export's own
            # claim, stored beside our sum so a disagreement is visible.
            if str(row[0] or "").strip().lower() == "total" and reported_net is None:
                raw = _cell(row, at, "net")
                if raw is not None and str(raw).strip() != "":
                    try:
                        reported_net = to_paise(raw)
                    except (ValueError, ArithmeticError):
                        warnings.append({"kind": "bad_total_row", "row": index})
            continue

        def cell(key: str, _row: tuple[Any, ...] = row) -> Any:
            return _cell(_row, at, key)

        bill_no = str(cell("bill_no") or "").strip()
        if not bill_no:
            continue

        status = str(cell("status") or "success").strip().lower()
        if status not in SUCCESS_STATUSES:
            warnings.append(
                {"kind": "skipped_status", "row": index, "bill_no": bill_no, "status": status}
            )
            continue

        try:
            ordered_at = parse_timestamp(cell("ordered_at"))
        except ValueError as exc:
            warnings.append(
                {"kind": "bad_timestamp", "row": index, "bill_no": bill_no, "detail": str(exc)}
            )
            continue

        if bill_no in seen:
            warnings.append({"kind": "duplicate_bill_no", "row": index, "bill_no": bill_no})
            continue
        seen.add(bill_no)

        raw_type = str(cell("order_type") or "").strip().lower()
        channel = CHANNELS.get(raw_type)
        if raw_type and channel is None:
            warnings.append({"kind": "unknown_order_type", "row": index, "value": raw_type})

        phone_hash = None
        raw_phone = cell("phone")
        if raw_phone not in (None, "") and hash_phone is not None:
            digits = re.sub(r"\D", "", str(raw_phone))
            if digits:
                phone_hash = hash_phone(digits)

        orders.append(
            ParsedOrder(
                external_bill_no=bill_no,
                ordered_at=ordered_at,
                # Never the calendar date in the cell: a bill at 00:05 belongs
                # to the night before.
                business_date=to_business_date(ordered_at),
                channel=channel,
                covers=_covers(cell("covers")),
                gross_paise=to_paise(cell("gross")),
                discount_paise=to_paise(cell("discount")),
                tax_paise=to_paise(cell("tax")),
                net_paise=to_paise(cell("net")),
                payment_mode=str(cell("payment") or "").strip() or None,
                table_no=str(cell("area") or "").strip() or None,
                customer_phone_hash=phone_hash,
            )
        )

    if not orders:
        raise UnreadableExport("The export contained no bills.")

    return ParseResult(
        orders=orders,
        warnings=warnings,
        period_start=period_start,
        period_end=period_end,
        restaurant=restaurant,
        reported_net_paise=reported_net,
    )
