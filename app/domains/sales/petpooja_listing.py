"""Reading a Petpooja Order Listing report.

The Orders Master Report carries every bill but no line items; this export
carries the item NAMES on each bill — comma-joined in a single `Items` cell —
and nothing else at line level: no quantity, no unit price. So this adapter
returns names per bill and leaves quantity honestly unknown, per the
open-items note that scoped it. The bills themselves come from the master;
this file only decorates them.

**What this export actually looks like**, established by reading AKIRA's own
file rather than from documentation:

    r0-r2   preamble — "Name:", "Restaurant Name:", "Restaurant Address:"
    r3      blank
    r4      the header row, 23 columns
    r5+     one row per order, newest first

The trap here is not summary rows — this report has none — but **split
payments**: a bill paid part-UPI, part-cash repeats its Order No. in extra
rows that carry only Grand Total and Payment Type. A parser keyed on Order
No. alone would either ingest bill 476 four times or report three duplicates
per split. A split row has no `Created` timestamp and no `Items`; that is the
discriminator.

Two more things the file taught:

- **`Order No.` is the master's `Invoice No.`** — verified cell-by-cell
  against the ingested master: all 89 orders match, amounts to the paisa. The
  write path joins on it and reports any order the master has not seen.
- **The file carries customer names, phones and addresses in the clear.**
  None of them leave the parse loop — not even hashed. The master already
  owns customer identity; this adapter's job is items, and the fewer places
  personal data passes through, the fewer places it can leak from.
"""

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

from app.core.business_date import business_date as to_business_date
from app.domains.sales.petpooja import (
    UnreadableExport,
    _cell,
    parse_timestamp,
    to_paise,
)

#: Bumped when this adapter's understanding of the export changes.
LISTING_ADAPTER_VERSION = "petpooja.listing.v1"

#: Identify the header row — searched for, not assumed at an offset. The
#: master is told apart by its own marker ("Invoice No."), which this report
#: does not have; requiring "Items" too keeps a third report shape from
#: slipping in on the strength of one column name.
LISTING_HEADER_MARKERS = ("Order No.", "Items")

COLUMNS = {
    "order_no": "Order No.",
    "items": "Items",
    "created": "Created",
    "amount": "My Amount (₹)",
    "status": "Status",
}
#: Without these the file is not an Order Listing report and parsing stops.
REQUIRED = ("order_no", "items")


@dataclass(frozen=True)
class ParsedListingOrder:
    external_bill_no: str
    #: Item names in bill order. Quantity is NOT known — the export does not
    #: carry it — which is why this is a list of names and nothing more.
    item_names: list[str]
    #: The Items cell verbatim, kept beside the split the same way raw_qty
    #: sits beside qty on a stock count line: the derivation stays checkable.
    raw_items: str
    #: When the order was struck, for the upload's period display only. The
    #: authoritative business_date comes from the matched master bill.
    created_at: datetime | None
    #: My Amount in paise — the master's net for the same bill, kept so the
    #: reconciliation is a stored column pair rather than a hand check.
    amount_paise: int | None


@dataclass(frozen=True)
class ListingResult:
    orders: list[ParsedListingOrder]
    warnings: list[dict[str, Any]] = field(default_factory=list)
    #: Split-payment continuation rows recognised and set aside. A count, not
    #: a warning: they are normal structure, and twelve warnings per file for
    #: normal structure would teach people to skip the warnings column.
    payment_split_rows: int = 0
    restaurant: str | None = None
    adapter_version: str = LISTING_ADAPTER_VERSION

    @property
    def period(self) -> tuple[date | None, date | None]:
        """Trading-day span of the parsed orders, from their Created stamps.
        This report has no period preamble, so the span is derived — and by
        business date, never the calendar date in the cell."""
        days = sorted(
            to_business_date(o.created_at) for o in self.orders if o.created_at is not None
        )
        if not days:
            return None, None
        return days[0], days[-1]

    @property
    def amount_paise(self) -> int:
        return sum(o.amount_paise or 0 for o in self.orders)


def split_items(raw: str) -> tuple[list[str], bool]:
    """The comma-joined Items cell as a list of names.

    Menu names contain parentheses — "Akira Shoyu Ramen (pork)" — and menu
    names are typed by humans into Petpooja, so one day a name will contain a
    comma inside its parentheses. A fragment with more "(" than ")" is
    re-joined with the next fragment; the flag reports whether that ever ran,
    so the row can carry a warning instead of a silently mangled name.
    """
    fragments = [f.strip() for f in raw.split(", ") if f.strip()]
    names: list[str] = []
    rejoined = False
    for fragment in fragments:
        if names and names[-1].count("(") > names[-1].count(")"):
            names[-1] = f"{names[-1]}, {fragment}"
            rejoined = True
        else:
            names.append(fragment)
    return names, rejoined


def parse_listing(data: bytes) -> ListingResult:
    """Read an Order Listing report. Bytes in, names-per-bill out."""
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise UnreadableExport(f"Not a readable .xlsx file: {exc}") from exc

    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    warnings: list[dict[str, Any]] = []
    restaurant = None
    header: list[str] | None = None
    header_at = -1

    for index, row in enumerate(rows):
        label = str(row[0] or "").strip().rstrip(":").lower() if row else ""
        if label == "restaurant name" and len(row) > 1:
            restaurant = str(row[1] or "").strip() or None
        cells = {str(cell or "").strip() for cell in row}
        if all(marker in cells for marker in LISTING_HEADER_MARKERS):
            header = [str(cell or "").strip() for cell in row]
            header_at = index
            break

    if header is None:
        raise UnreadableExport(
            "No header row containing 'Order No.' and 'Items'. "
            "This does not look like a Petpooja Order Listing report."
        )

    at: dict[str, int] = {}
    for key, name in COLUMNS.items():
        if name in header:
            at[key] = header.index(name)
        elif key in REQUIRED:
            raise UnreadableExport(f"The export has no {name!r} column.")
        else:
            warnings.append({"kind": "missing_column", "column": name, "effect": f"{key} not set"})

    orders: list[ParsedListingOrder] = []
    seen: set[str] = set()
    split_rows = 0

    for index, row in enumerate(rows[header_at + 1 :], start=header_at + 2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        def cell(key: str, _row: tuple[Any, ...] = row) -> Any:
            return _cell(_row, at, key)

        bill_no = str(cell("order_no") or "").strip()
        if not bill_no:
            continue

        raw_items = str(cell("items") or "").strip()
        created_raw = cell("created")
        if not raw_items and (created_raw is None or str(created_raw).strip() == ""):
            # A split-payment continuation: the same Order No. again, carrying
            # only a partial amount and its payment type. Normal structure —
            # counted, not warned about, and never a duplicate.
            split_rows += 1
            continue

        if bill_no in seen:
            warnings.append({"kind": "duplicate_order", "row": index, "bill_no": bill_no})
            continue
        seen.add(bill_no)

        if not raw_items:
            warnings.append({"kind": "empty_items", "row": index, "bill_no": bill_no})
            continue

        created_at: datetime | None = None
        if created_raw is not None and str(created_raw).strip():
            try:
                created_at = parse_timestamp(created_raw)
            except ValueError as exc:
                warnings.append(
                    {"kind": "bad_timestamp", "row": index, "bill_no": bill_no, "detail": str(exc)}
                )

        names, rejoined = split_items(raw_items)
        if rejoined:
            warnings.append({"kind": "rejoined_item_name", "row": index, "bill_no": bill_no})

        amount_raw = cell("amount")
        amount = None
        if amount_raw is not None and str(amount_raw).strip() != "":
            try:
                amount = to_paise(amount_raw)
            except (ValueError, ArithmeticError):
                warnings.append({"kind": "bad_amount", "row": index, "bill_no": bill_no})

        orders.append(
            ParsedListingOrder(
                external_bill_no=bill_no,
                item_names=names,
                raw_items=raw_items,
                created_at=created_at,
                amount_paise=amount,
            )
        )

    if not orders:
        raise UnreadableExport("The export contained no orders.")

    return ListingResult(
        orders=orders,
        warnings=warnings,
        payment_split_rows=split_rows,
        restaurant=restaurant,
    )
