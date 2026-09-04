"""The restaurant-name guard on sales uploads.

Every Petpooja report names its venue in the preamble. Until this guard, all
three adapters read that line and dropped it, so another restaurant's export
uploaded against an Akira outlet ingested silently and looked, afterwards,
like an unremarkable few weeks of trading.

Two properties are worth holding down with tests rather than reading:

- the guard is OFF until the setting has a value, so shipping it does not
  lock anyone out of uploading, and
- once ON it fails CLOSED — a file with no restaurant line is refused too,
  because "cannot check" is not "checked and fine".
"""

import json
import uuid
from io import BytesIO

import pytest
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine

from app.core.errors import ValidationError
from app.domains.sales import petpooja, service

pytestmark = pytest.mark.asyncio

KEY = "sales.petpooja_restaurant_name"


def master_export(restaurant: str | None = "Akira Ramen") -> bytes:
    """The Orders Master Report's shape, trimmed to what the guard reads."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Date:", "2026-08-01 to 2026-08-28"])
    ws.append(["Name:", "Orders Master Report"])
    if restaurant is not None:
        ws.append(["Restaurant Name:", restaurant])
    ws.append([])
    ws.append(["Invoice No.", "Date", "Net Sales (₹)(M.A - D)"])
    ws.append(["1", "2026-08-22 20:00:00", "410.0"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def listing_export(restaurant: str = "Akira Ramen") -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Name:", "Order Listing"])
    ws.append(["Restaurant Name:", restaurant])
    ws.append([])
    ws.append(["Order No.", "Items", "Created", "My Amount (₹)"])
    ws.append(["1", "Veg Ramen", "26 Aug 2026 00:46:15", "410.0"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def itemdays_export(restaurant: str = "Akira Ramen") -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Date:", "2026-08-01 to 2026-08-28"])
    ws.append(["Restaurant Name:", restaurant])
    ws.append([])
    ws.append(["Item", "Date", "Qty.", "Total (₹)"])
    ws.append(["Ramen", "2026-08-21", "5.0", "2100.0"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestNormalising:
    """Fold what two spellings of one venue share; fold nothing else."""

    @pytest.mark.parametrize(
        "a,b",
        [
            ("Akira Ramen", "AKIRA RAMEN"),
            ("Akira Ramen", "  Akira   Ramen  "),
            ("Akira Ramen", "akira ramen"),
            ("Akira Ramen", "Akira\tRamen"),
        ],
    )
    def test_case_and_whitespace_are_the_same_venue(self, a: str, b: str) -> None:
        assert petpooja.normalise_restaurant(a) == petpooja.normalise_restaurant(b)

    @pytest.mark.parametrize(
        "a,b",
        [
            # The whole point: a prefix must NOT match, or "Akira" would
            # accept every Akira-branded venue in the country.
            ("Akira", "Akira Ramen Bangalore"),
            ("Akira Ramen", "Akira Ramen Bangalore"),
            ("Akira Ramen", "Akira Sushi"),
            ("Akira Ramen", "Akira-Ramen"),
        ],
    )
    def test_different_venues_stay_different(self, a: str, b: str) -> None:
        assert petpooja.normalise_restaurant(a) != petpooja.normalise_restaurant(b)

    def test_nothing_normalises_to_the_same_thing_as_empty(self) -> None:
        assert petpooja.normalise_restaurant(None) == ""
        assert petpooja.normalise_restaurant("   ") == ""
        assert petpooja.normalise_restaurant("Akira") != ""


class TestReadingThePreamble:
    """The name comes back verbatim, from every report shape."""

    def test_all_three_reports_are_read(self) -> None:
        assert service.inspect_export(master_export()).restaurant == "Akira Ramen"
        assert service.inspect_export(listing_export()).restaurant == "Akira Ramen"
        assert service.inspect_export(itemdays_export()).restaurant == "Akira Ramen"

    def test_the_report_type_is_still_detected(self) -> None:
        """The guard rides along with detection; it must not disturb it."""
        assert service.detect_source(master_export()) == service.SOURCE_ORDERS
        assert service.detect_source(listing_export()) == service.SOURCE_LISTING
        assert service.detect_source(itemdays_export()) == service.SOURCE_ITEMDAYS

    def test_a_file_without_the_line_reads_as_none_not_empty(self) -> None:
        """None means "the file did not say"; "" would be a name."""
        assert service.inspect_export(master_export(restaurant=None)).restaurant is None

    def test_it_is_kept_verbatim_not_normalised(self) -> None:
        """What is stored must be what the file said — that is the string an
        admin copies into the setting."""
        found = service.inspect_export(master_export("  AKIRA  Ramen ")).restaurant
        assert found == "AKIRA  Ramen"


@pytest.fixture
async def session(migrated_db: str):  # type: ignore[no-untyped-def]
    engine = create_async_engine(migrated_db.replace("postgresql://", "postgresql+asyncpg://"))
    async with AsyncSession(engine, expire_on_commit=False) as db:
        try:
            yield db
        finally:
            await db.rollback()
            await db.execute(text("delete from app_settings where key = :k"), {"k": KEY})
            await db.commit()
    await engine.dispose()


async def _outlets(db: AsyncSession) -> list[uuid.UUID]:
    rows = (await db.execute(text("select id from outlets order by code"))).scalars().all()
    return [uuid.UUID(str(r)) for r in rows]


async def _arm(db: AsyncSession, value: str, *, outlet_id: uuid.UUID | None = None) -> None:
    """Set the expected name the way the admin screen would."""
    await db.execute(
        text(
            """
            insert into app_settings (key, scope, outlet_id, value, effective_from)
            values (:k, cast(:scope as setting_scope), :o, cast(:v as jsonb),
                    now() - interval '1 minute')
            """
        ),
        {
            "k": KEY,
            "scope": "outlet" if outlet_id else "global",
            "o": outlet_id,
            "v": json.dumps(value),
        },
    )
    await db.commit()


class TestTheGuard:
    async def test_it_is_off_until_the_setting_has_a_value(self, session: AsyncSession) -> None:
        """Shipping this must not stop uploads at outlets nobody has
        configured yet. The default is empty, and empty means accept."""
        outlet = (await _outlets(session))[0]
        await service.check_restaurant(session, outlet_id=outlet, found="Somebody Else Entirely")

    async def test_the_expected_name_passes(self, session: AsyncSession) -> None:
        outlet = (await _outlets(session))[0]
        await _arm(session, "Akira Ramen")
        await service.check_restaurant(session, outlet_id=outlet, found="Akira Ramen")

    async def test_another_venue_is_refused(self, session: AsyncSession) -> None:
        outlet = (await _outlets(session))[0]
        await _arm(session, "Akira Ramen")
        with pytest.raises(ValidationError) as caught:
            await service.check_restaurant(session, outlet_id=outlet, found="Sushi Bar Salt Lake")
        # Both names in the message: the fix is either the setting or the
        # file, and the reader cannot tell which without seeing both.
        assert "Sushi Bar Salt Lake" in str(caught.value)
        assert "Akira Ramen" in str(caught.value)

    async def test_a_differently_written_same_name_passes(self, session: AsyncSession) -> None:
        outlet = (await _outlets(session))[0]
        await _arm(session, "Akira Ramen")
        await service.check_restaurant(session, outlet_id=outlet, found="  AKIRA   ramen ")

    async def test_a_file_with_no_name_is_refused_once_armed(self, session: AsyncSession) -> None:
        """Fails closed. "Cannot be checked" is not "checked and fine" — a
        stripped preamble must not be the way in."""
        outlet = (await _outlets(session))[0]
        await _arm(session, "Akira Ramen")
        with pytest.raises(ValidationError) as caught:
            await service.check_restaurant(session, outlet_id=outlet, found=None)
        assert "Restaurant Name:" in str(caught.value)

    async def test_an_outlet_override_beats_the_global_value(self, session: AsyncSession) -> None:
        """One group, two Petpooja accounts is a real shape — D9's override
        mechanism is why this needed no new table."""
        outlets = await _outlets(session)
        await _arm(session, "Akira Ramen")
        await _arm(session, "Akira Ramen Express", outlet_id=outlets[1])

        await service.check_restaurant(session, outlet_id=outlets[0], found="Akira Ramen")
        await service.check_restaurant(session, outlet_id=outlets[1], found="Akira Ramen Express")
        with pytest.raises(ValidationError):
            await service.check_restaurant(session, outlet_id=outlets[1], found="Akira Ramen")


class TestTheSettingIsAdministrable:
    """It has to be changeable from the admin screen, or arming it means a
    deploy — which is exactly what D9 exists to avoid."""

    def test_it_is_declared_in_the_registry(self) -> None:
        from app.core.settings_registry import REGISTRY, validate_value

        definition = REGISTRY[KEY]
        assert definition.group == "sales"
        assert definition.type == "string"
        assert definition.default == ""
        assert definition.outlet_overridable is True
        # Free text, not a closed set: nobody can enumerate restaurant names.
        assert definition.choices == ()
        assert validate_value(definition, "Akira Ramen") is None
        assert validate_value(definition, 7) is not None


class TestWhatIsRecorded:
    async def test_the_claimed_name_is_stored_on_the_upload(self, session: AsyncSession) -> None:
        """Recorded whether or not the guard was armed — that is what lets
        "has anything foreign ever landed here" be a query over the whole
        history, and what an admin copies into the setting."""
        outlet = (await _outlets(session))[0]
        upload_id = (
            await session.execute(
                text(
                    """
                    insert into data_uploads
                        (outlet_id, source, original_filename, storage_path,
                         file_sha256, status, restaurant_name)
                    values (:o, 'petpooja_orders', 'e.xlsx', :p, :sha,
                            'received', :r)
                    returning id
                    """
                ),
                {
                    "o": outlet,
                    "p": f"{outlet}/guardtest.xlsx",
                    "sha": "guardtest-sha",
                    "r": "Akira Ramen",
                },
            )
        ).scalar_one()
        await session.commit()
        try:
            stored = (
                await session.execute(
                    text("select restaurant_name from data_uploads where id = :i"),
                    {"i": upload_id},
                )
            ).scalar_one()
            assert stored == "Akira Ramen"
        finally:
            await session.execute(text("delete from data_uploads where id = :i"), {"i": upload_id})
            await session.commit()
